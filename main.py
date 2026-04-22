from fastapi import FastAPI, Form, File, UploadFile, HTTPException, Body
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from uuid import uuid4
from io import BytesIO
from typing import Any

import json
import os
import shutil
import tempfile
import traceback
import zipfile
import requests

from PIL import Image


app = FastAPI()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_ROLE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
SUPABASE_BUCKET = "excel-image-formatted-api"

IMAGE_KEY = "product_image"

# Faster settings
IMAGE_EXPORT_SIZE = 180
DISPLAY_IMAGE_WIDTH = 80
DISPLAY_IMAGE_HEIGHT = 80

ROW_HEIGHT = 62
IMAGE_COL_WIDTH = 12

DEFAULT_COL_WIDTH = 16
TITLE_COL_WIDTH = 30
SKU_COL_WIDTH = 24
UPC_COL_WIDTH = 18

SESSION = requests.Session()


@app.get("/")
def root():
    return {"ok": True}


@app.get("/health")
def health():
    return {"ok": True}


@app.post("/export")
async def export_excel(
    payload: str = Form(...),
    image_manifest: str = Form(...),
    images_zip: UploadFile = File(...),
):
    temp_dir = tempfile.mkdtemp(prefix="excel_export_")

    try:
        raw = json.loads(payload)
        manifest = json.loads(image_manifest)

        request_body = raw.get("body", raw)
        rows = request_body.get("data")

        if not isinstance(rows, list) or not rows:
            raise HTTPException(status_code=400, detail='Request must contain "data" as a non-empty array.')

        if not isinstance(manifest, dict):
            raise HTTPException(status_code=400, detail="image_manifest must be a JSON object.")

        zip_path = os.path.join(temp_dir, images_zip.filename or "images.zip")

        with open(zip_path, "wb") as f:
            shutil.copyfileobj(images_zip.file, f)

        extracted_dir = os.path.join(temp_dir, "images")
        os.makedirs(extracted_dir, exist_ok=True)

        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            zip_ref.extractall(extracted_dir)

        image_path_map = {}
        for image_url, filename in manifest.items():
            local_path = os.path.join(extracted_dir, filename)
            if os.path.exists(local_path):
                processed_path = make_square_fill_image(local_path, temp_dir, filename)
                image_path_map[image_url] = processed_path

        public_url = process_export_job(rows, image_path_map, temp_dir)

        return JSONResponse({
            "ok": True,
            "url": public_url,
        })

    except HTTPException:
        raise
    except Exception as e:
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cleanup_temp_dir(temp_dir)


def process_export_job(rows: list[dict], image_path_map: dict[str, str], temp_dir: str) -> str:
    print("JOB START")
    print(f"Rows received: {len(rows)}")

    headers = collect_headers(rows)
    print(f"Headers: {headers}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.freeze_panes = "A2"

    write_headers(ws, headers)
    write_rows(ws, rows, headers)
    set_fixed_column_widths(ws, headers)
    embed_images_from_zip(ws, rows, headers, image_path_map)

    file_uuid = str(uuid4())
    filename = f"{file_uuid}.xlsx"
    output_path = os.path.join(temp_dir, filename)

    print("Saving workbook...")
    wb.save(output_path)
    print(f"Workbook saved: {output_path}")

    print("Uploading to Supabase...")
    public_url = upload_to_supabase(output_path, filename)
    print(f"Upload complete: {filename}")
    print("JOB END")

    return public_url


def collect_headers(rows: list[dict]) -> list[str]:
    headers = []
    seen = set()

    for row in rows:
        if not isinstance(row, dict):
            continue
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)

    return headers


def write_headers(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = Font(bold=True)


def write_rows(ws, rows: list[dict], headers: list[str]):
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=str(value))


def set_fixed_column_widths(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        header_lower = str(header).strip().lower()

        if header == IMAGE_KEY:
            ws.column_dimensions[col_letter].width = IMAGE_COL_WIDTH
        elif "title" in header_lower or "name" in header_lower:
            ws.column_dimensions[col_letter].width = TITLE_COL_WIDTH
        elif "sku" in header_lower:
            ws.column_dimensions[col_letter].width = SKU_COL_WIDTH
        elif "upc" in header_lower or "barcode" in header_lower:
            ws.column_dimensions[col_letter].width = UPC_COL_WIDTH
        else:
            ws.column_dimensions[col_letter].width = DEFAULT_COL_WIDTH


def make_square_fill_image(source_path: str, temp_dir: str, filename: str) -> str:
    out_path = os.path.join(temp_dir, f"processed_{os.path.splitext(filename)[0]}.jpg")

    with Image.open(source_path) as img:
        img = img.convert("RGB")
        width, height = img.size

        if width > height:
            offset = (width - height) // 2
            img = img.crop((offset, 0, offset + height, height))
        elif height > width:
            offset = (height - width) // 2
            img = img.crop((0, offset, width, offset + width))

        img = img.resize((IMAGE_EXPORT_SIZE, IMAGE_EXPORT_SIZE), Image.BILINEAR)
        img.save(out_path, format="JPEG", quality=85, optimize=True)

    return out_path


def embed_images_from_zip(ws, rows: list[dict], headers: list[str], image_path_map: dict[str, str]):
    if IMAGE_KEY not in headers:
        print("No product_image column found, skipping image embedding.")
        return

    image_col_idx = headers.index(IMAGE_KEY) + 1
    col_letter = get_column_letter(image_col_idx)

    for row_idx, row in enumerate(rows, start=2):
        image_url = row.get(IMAGE_KEY, "")
        image_url = "" if image_url is None else str(image_url).strip()

        if not image_url:
            continue

        img_path = image_path_map.get(image_url)
        if not img_path or not os.path.exists(img_path):
            continue

        try:
            ws.cell(row=row_idx, column=image_col_idx, value="")
            ws.row_dimensions[row_idx].height = ROW_HEIGHT

            xl_img = XLImage(img_path)
            xl_img.width = DISPLAY_IMAGE_WIDTH
            xl_img.height = DISPLAY_IMAGE_HEIGHT
            xl_img.anchor = f"{col_letter}{row_idx}"
            ws.add_image(xl_img)

        except Exception as e:
            print(f"Image embed failed for row {row_idx}: {e}")
            ws.cell(row=row_idx, column=image_col_idx, value=image_url)


def upload_to_supabase(file_path: str, filename: str) -> str:
    upload_url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{filename}"

    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "x-upsert": "false",
    }

    with open(file_path, "rb") as f:
        response = SESSION.post(upload_url, headers=headers, data=f, timeout=(5, 120))

    if response.status_code not in (200, 201):
        raise Exception(f"Supabase upload failed: {response.status_code} {response.text}")

    return f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{filename}"


def cleanup_temp_dir(temp_dir: str):
    try:
        if os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception:
        pass

def sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    safe = str(name or "Sheet")
    for char in ['\\', '/', '?', '*', '[', ']', ':']:
        safe = safe.replace(char, "")
    safe = safe.strip() or "Sheet"
    safe = safe[:31]

    final_name = safe
    counter = 1

    while final_name in used_names:
        suffix = f"_{counter}"
        final_name = safe[: 31 - len(suffix)] + suffix
        counter += 1

    used_names.add(final_name)
    return final_name


@app.post("/create-updated-orders-xlsx")
async def create_updated_orders_xlsx(data: list[dict[str, Any]] = Body(...)):
    if not isinstance(data, list) or not data:
        raise HTTPException(status_code=400, detail="Request body must be a non-empty array.")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_names: set[str] = set()

    export_headers = [
        "PO Number",
        "PO Line #",
        "Qty Ordered",
        "Unit Price",
        "Buyers Catalog or Stock Keeping #",
        "UPC/EAN",
        "Vendor Style",
        "Retail Price",
        "Product/Item Description",
        "updated_quantity",
        "cut_applied",
    ]

    totals_by_sku: dict[str, dict[str, Any]] = {}
    sheets_created = 0

    def to_number(value: Any) -> float:
        try:
            if value is None or str(value).strip() == "":
                return 0
            return float(str(value).strip())
        except Exception:
            return 0

    for order in data:
        nested_items = order.get("items") or []
        updated_rows = [
            row for row in nested_items
            if isinstance(row, dict) and row.get("updated") is True
        ]

        if not updated_rows:
            continue

        po_number = order.get("PO Number", "Sheet")
        sheet_name = sanitize_sheet_name(po_number, used_names)
        ws = wb.create_sheet(title=sheet_name)
        sheets_created += 1

        ws.append(export_headers)

        for row in updated_rows:
            ws.append([
                row.get("PO Number", ""),
                row.get("PO Line #", ""),
                row.get("Qty Ordered", ""),
                row.get("Unit Price", ""),
                row.get("Buyers Catalog or Stock Keeping #", ""),
                row.get("UPC/EAN", ""),
                row.get("Vendor Style", ""),
                row.get("Retail Price", ""),
                row.get("Product/Item Description", ""),
                row.get("updated_quantity", ""),
                row.get("cut_applied", ""),
            ])

            sku = str(row.get("Vendor Style", "")).strip()
            if not sku:
                sku = str(row.get("Buyers Catalog or Stock Keeping #", "")).strip()
            if not sku:
                sku = "UNKNOWN"

            if sku not in totals_by_sku:
                totals_by_sku[sku] = {
                    "Vendor Style": sku,
                    "UPC/EAN": row.get("UPC/EAN", ""),
                    "Product/Item Description": row.get("Product/Item Description", ""),
                    "total_original_ordered": 0,
                    "total_cut": 0,
                }

            totals_by_sku[sku]["total_original_ordered"] += to_number(row.get("Qty Ordered", 0))
            totals_by_sku[sku]["total_cut"] += to_number(row.get("cut_applied", 0))

    totals_ws = wb.create_sheet(title="Totals", index=0)
    totals_headers = [
        "Vendor Style",
        "UPC/EAN",
        "Product/Item Description",
        "total_original_ordered",
        "total_cut",
    ]
    totals_ws.append(totals_headers)

    for sku in sorted(totals_by_sku.keys()):
        entry = totals_by_sku[sku]
        totals_ws.append([
            entry["Vendor Style"],
            entry["UPC/EAN"],
            entry["Product/Item Description"],
            entry["total_original_ordered"],
            entry["total_cut"],
        ])

    if sheets_created == 0:
        no_updates_ws = wb.create_sheet(title="No Updates")
        no_updates_ws.append(["No updated items found"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="golf_town_updated_orders.xlsx"'
        },
    )
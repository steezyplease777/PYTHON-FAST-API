from io import BytesIO
from typing import Any

from fastapi import APIRouter, Body, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from app.utils.formatting import sanitize_sheet_name

router = APIRouter()


@router.post("/create-updated-orders-xlsx")
async def create_updated_orders_xlsx(data: list[dict[str, Any]] = Body(...)):
    if not isinstance(data, list) or not data:
        raise HTTPException(status_code=400, detail="Request body must be a non-empty array.")

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_names: set[str] = set()

    export_headers = [
        "PO Number",
        "PO Line #",
        "Qty Ordered",
        "Unit Price",
        "Buyers Catalog or Stock Keeping #",
        "UPC/EAN",
        "Vendor Style",
        "Retail Price",
        "Product/Item Description",
        "updated_quantity",
        "cut_applied",
    ]

    totals_by_sku: dict[str, dict[str, Any]] = {}
    sheets_created = 0

    def to_number(value: Any) -> float:
        try:
            if value is None or str(value).strip() == "":
                return 0
            return float(str(value).strip())
        except Exception:
            return 0

    for order in data:
        nested_items = order.get("items") or []
        updated_rows = [
            row for row in nested_items
            if isinstance(row, dict) and row.get("updated") is True
        ]

        if not updated_rows:
            continue

        po_number = order.get("PO Number", "Sheet")
        sheet_name = sanitize_sheet_name(po_number, used_names)
        ws = wb.create_sheet(title=sheet_name)
        sheets_created += 1

        ws.append(export_headers)

        for row in updated_rows:
            ws.append([
                row.get("PO Number", ""),
                row.get("PO Line #", ""),
                row.get("Qty Ordered", ""),
                row.get("Unit Price", ""),
                row.get("Buyers Catalog or Stock Keeping #", ""),
                row.get("UPC/EAN", ""),
                row.get("Vendor Style", ""),
                row.get("Retail Price", ""),
                row.get("Product/Item Description", ""),
                row.get("updated_quantity", ""),
                row.get("cut_applied", ""),
            ])

            sku = str(row.get("Vendor Style", "")).strip()
            if not sku:
                sku = str(row.get("Buyers Catalog or Stock Keeping #", "")).strip()
            if not sku:
                sku = "UNKNOWN"

            if sku not in totals_by_sku:
                totals_by_sku[sku] = {
                    "Vendor Style": sku,
                    "UPC/EAN": row.get("UPC/EAN", ""),
                    "Product/Item Description": row.get("Product/Item Description", ""),
                    "total_original_ordered": 0,
                    "total_cut": 0,
                }

            totals_by_sku[sku]["total_original_ordered"] += to_number(row.get("Qty Ordered", 0))
            totals_by_sku[sku]["total_cut"] += to_number(row.get("cut_applied", 0))

    totals_ws = wb.create_sheet(title="Totals", index=0)
    totals_headers = [
        "Vendor Style",
        "UPC/EAN",
        "Product/Item Description",
        "total_original_ordered",
        "total_cut",
    ]
    totals_ws.append(totals_headers)

    for sku in sorted(totals_by_sku.keys()):
        entry = totals_by_sku[sku]
        totals_ws.append([
            entry["Vendor Style"],
            entry["UPC/EAN"],
            entry["Product/Item Description"],
            entry["total_original_ordered"],
            entry["total_cut"],
        ])

    if sheets_created == 0:
        no_updates_ws = wb.create_sheet(title="No Updates")
        no_updates_ws.append(["No updated items found"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="golf_town_updated_orders.xlsx"'
        },
    )

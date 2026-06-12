import os
import shutil
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from PIL import Image

from app.config import (
    IMAGE_KEY,
    IMAGE_EXPORT_SIZE,
    DISPLAY_IMAGE_WIDTH,
    DISPLAY_IMAGE_HEIGHT,
    ROW_HEIGHT,
    IMAGE_COL_WIDTH,
    DEFAULT_COL_WIDTH,
    TITLE_COL_WIDTH,
    SKU_COL_WIDTH,
    UPC_COL_WIDTH,
)
from app.services.storage_service import upload_xlsx_to_supabase


def process_export_job(rows: list[dict], image_path_map: dict[str, str], temp_dir: str) -> str:
    print("JOB START")
    print(f"Rows received: {len(rows)}")

    headers = collect_headers(rows)
    print(f"Headers: {headers}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.freeze_panes = "A2"

    write_headers(ws, headers)
    write_rows(ws, rows, headers)
    set_fixed_column_widths(ws, headers)
    embed_images_from_zip(ws, rows, headers, image_path_map)

    file_uuid = str(uuid4())
    filename = f"{file_uuid}.xlsx"
    output_path = os.path.join(temp_dir, filename)

    print("Saving workbook...")
    wb.save(output_path)
    print(f"Workbook saved: {output_path}")

    print("Uploading to Supabase...")
    public_url = upload_xlsx_to_supabase(output_path, filename)
    print(f"Upload complete: {filename}")
    print("JOB END")

    return public_url


def collect_headers(rows: list[dict]) -> list[str]:
    headers = []
    seen = set()

    for row in rows:
        if not isinstance(row, dict):
            continue
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)

    return headers


def write_headers(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(header))
        cell.font = Font(bold=True)


def write_rows(ws, rows: list[dict], headers: list[str]):
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            if value is None:
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=str(value))


def set_fixed_column_widths(ws, headers: list[str]):
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        header_lower = str(header).strip().lower()

        if header == IMAGE_KEY:
            ws.column_dimensions[col_letter].width = IMAGE_COL_WIDTH
        elif "title" in header_lower or "name" in header_lower:
            ws.column_dimensions[col_letter].width = TITLE_COL_WIDTH
        elif "sku" in header_lower:
            ws.column_dimensions[col_letter].width = SKU_COL_WIDTH
        elif "upc" in header_lower or "barcode" in header_lower:
            ws.column_dimensions[col_letter].width = UPC_COL_WIDTH
        else:
            ws.column_dimensions[col_letter].width = DEFAULT_COL_WIDTH


def make_square_fill_image(source_path: str, temp_dir: str, filename: str) -> str:
    out_path = os.path.join(temp_dir, f"processed_{os.path.splitext(filename)[0]}.jpg")

    with Image.open(source_path) as img:
        img = img.convert("RGB")
        width, height = img.size

        if width > height:
            offset = (width - height) // 2
            img = img.crop((offset, 0, offset + height, height))
        elif height > width:
            offset = (height - width) // 2
            img = img.crop((0, offset, width, offset + width))

        img = img.resize((IMAGE_EXPORT_SIZE, IMAGE_EXPORT_SIZE), Image.BILINEAR)
        img.save(out_path, format="JPEG", quality=85, optimize=True)

    return out_path


def embed_images_from_zip(ws, rows: list[dict], headers: list[str], image_path_map: dict[str, str]):
    if IMAGE_KEY not in headers:
        print("No product_image column found, skipping image embedding.")
        return

    image_col_idx = headers.index(IMAGE_KEY) + 1
    col_letter = get_column_letter(image_col_idx)

    for row_idx, row in enumerate(rows, start=2):
        image_url = row.get(IMAGE_KEY, "")
        image_url = "" if image_url is None else str(image_url).strip()

        if not image_url:
            continue

        img_path = image_path_map.get(image_url)
        if not img_path or not os.path.exists(img_path):
            continue

        try:
            ws.cell(row=row_idx, column=image_col_idx, value="")
            ws.row_dimensions[row_idx].height = ROW_HEIGHT

            xl_img = XLImage(img_path)
            xl_img.width = DISPLAY_IMAGE_WIDTH
            xl_img.height = DISPLAY_IMAGE_HEIGHT
            xl_img.anchor = f"{col_letter}{row_idx}"
            ws.add_image(xl_img)

        except Exception as e:
            print(f"Image embed failed for row {row_idx}: {e}")
            ws.cell(row=row_idx, column=image_col_idx, value=image_url)


def cleanup_temp_dir(temp_dir: str):
    try:
        if os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception:
        pass

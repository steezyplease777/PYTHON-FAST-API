from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import PlainTextResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from PIL import Image
from io import BytesIO
from uuid import uuid4

import os
import tempfile
import requests


app = FastAPI()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_ROLE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
SUPABASE_BUCKET = "excel-image-formatted-api"

IMAGE_KEY = "product_image"
IMAGE_WIDTH = 80
IMAGE_HEIGHT = 80
ROW_HEIGHT = 65
REQUEST_TIMEOUT = 20


@app.get("/")
def root():
    return {"ok": True}


@app.post("/export", response_class=PlainTextResponse)
async def export_excel(request: Request):
    try:
        body = await request.json()
        rows = body.get("data")

        if not isinstance(rows, list) or len(rows) == 0:
            raise HTTPException(status_code=400, detail='Body must contain "data" as a non-empty array.')

        # collect all headers from all objects, preserving first-seen order
        headers = []
        seen = set()
        for row in rows:
            if not isinstance(row, dict):
                raise HTTPException(status_code=400, detail='Each item in "data" must be a JSON object.')
            for key in row.keys():
                if key not in seen:
                    seen.add(key)
                    headers.append(key)

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"

        # write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"

        # write data as text first
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row.get(header, "")
                if value is None:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=str(value))

        # set widths
        for col_idx, header in enumerate(headers, start=1):
            if header == IMAGE_KEY:
                ws.column_dimensions[get_column_letter(col_idx)].width = 14
            else:
                longest = len(header)
                for row in rows[:500]:
                    value = row.get(header, "")
                    text = "" if value is None else str(value)
                    if len(text) > longest:
                        longest = len(text)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(longest + 2, 10), 40)

        # embed images
        image_temp_paths = []

        if IMAGE_KEY in headers:
            image_col_idx = headers.index(IMAGE_KEY) + 1

            for row_idx, row in enumerate(rows, start=2):
                image_url = row.get(IMAGE_KEY, "")
                image_url = "" if image_url is None else str(image_url).strip()

                if not image_url:
                    continue

                try:
                    img_path = download_and_prepare_image(image_url, row_idx)
                    image_temp_paths.append(img_path)

                    # remove url text from the cell
                    ws.cell(row=row_idx, column=image_col_idx, value="")
                    ws.row_dimensions[row_idx].height = ROW_HEIGHT

                    xl_img = XLImage(img_path)
                    xl_img.width = IMAGE_WIDTH
                    xl_img.height = IMAGE_HEIGHT
                    xl_img.anchor = f"{get_column_letter(image_col_idx)}{row_idx}"
                    ws.add_image(xl_img)
                except Exception:
                    # if image fails, leave original URL text in the cell
                    ws.cell(row=row_idx, column=image_col_idx, value=image_url)

        # save workbook
        file_uuid = str(uuid4())
        filename = f"{file_uuid}.xlsx"

        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, filename)
        wb.save(output_path)

        # upload to supabase storage
        upload_to_supabase(output_path, filename)

        return PlainTextResponse("ok", status_code=200)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def download_and_prepare_image(url: str, row_idx: int) -> str:
    response = requests.get(url, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()

    img = Image.open(BytesIO(response.content))

    # normalize to png for openpyxl compatibility
    if img.mode not in ("RGB", "RGBA"):
        img = img.convert("RGBA")

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, f"product_image_{row_idx}.png")
    img.save(file_path, format="PNG")

    return file_path


def upload_to_supabase(file_path: str, filename: str):
    upload_url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{filename}"

    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "x-upsert": "false",
    }

    with open(file_path, "rb") as f:
        response = requests.post(upload_url, headers=headers, data=f)

    if response.status_code not in (200, 201):
        raise Exception(f"Supabase upload failed: {response.status_code} {response.text}")